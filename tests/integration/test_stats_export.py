"""Integration tests for the tenant statistics / runs-explorer export (task_14_14).

The export surface behind ``GET /tenant-stats/runs/export``: the same
tenant-scoped, filterable runs-explorer rows serialised to a downloadable file
in one of three formats:

  * ``csv``  — stdlib :mod:`csv`, the lowest-common-denominator format;
  * ``xlsx`` — :mod:`openpyxl` (pure-Python, pip-clean) — the native
    spreadsheet format the runs explorer calls for (Resumen, section 13.8);
  * ``pdf``  — DEGRADED to a print-ready ``text/html`` document (the
    api-server image ships no native PDF renderer — same call as the
    docs-viewer PDF export — so the browser's "Save as PDF" closes the loop).

Like the dashboard endpoints it is JWT-authenticated, gated on ``tenant_admin``
and runs on a tenant-scoped RLS session. Costs are canonical USD.

Coverage:

  * CSV: right header + one row per run with the resolved labels / model /
    verdict / cost for the tenant's data;
  * XLSX: a valid ``.xlsx`` (PK zip magic) that opens with openpyxl, with the
    same header + a row per run;
  * PDF: a valid (well-formed) HTML document carrying the runs;
  * the export is tenant-scoped (@pytest.mark.cross_tenant): tenant A's export
    never contains tenant B's rows;
  * RBAC: a plain member (tenant_user) is denied (403);
  * filters narrow the export the same way as the JSON explorer;
  * no secret / PII (steps_log payload, prompts) leaks into the export.

Pre-condition: postgres (15432) + redis (6379) from docker-compose are healthy;
the fixtures create a throwaway DB and flush Redis DB 15.
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
from datetime import UTC, datetime, timedelta
from uuid import UUID, uuid4

import asyncpg
import pytest
from alembic import command
from api_server.db.execution_repo import steps_rollup
from httpx import ASGITransport, AsyncClient

from ._partitions import ensure_partition_for

pytestmark = pytest.mark.integration


# ---------------------------------------------------------------------------
# DB seed helpers (BYPASSRLS via migrations_user DSN)
# ---------------------------------------------------------------------------
async def _seed_tenant(dsn: str, *, slug: str) -> UUID:
    tenant = uuid4()
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO organizations (id, name, slug) VALUES ($1, $2, $3)",
            tenant,
            slug.title(),
            slug,
        )
    finally:
        await conn.close()
    return tenant


async def _seed_user_with_jwt(
    dsn: str, redis_url: str, *, tenant_id: UUID, email: str, role: str
) -> tuple[UUID, str]:
    from api_server.auth.jwt import encode_jwt
    from api_server.auth.sessions import SessionStore
    from redis.asyncio import Redis
    from uuid6 import uuid7

    user_id = uuid4()
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO users (id, email, password_hash, is_system_admin) "
            "VALUES ($1, $2, $3, false)",
            user_id,
            email,
            "x",
        )
        await conn.execute(
            "INSERT INTO user_org_memberships (id, tenant_id, user_id, role, is_active) "
            "VALUES ($1, $2, $3, $4, true)",
            uuid4(),
            tenant_id,
            user_id,
            role,
        )
    finally:
        await conn.close()

    session_id = uuid7()
    redis: Redis = Redis.from_url(redis_url, decode_responses=True)
    try:
        await SessionStore(redis).create(
            session_id, user_id=user_id, tenant_id=tenant_id, ttl_seconds=3600
        )
    finally:
        await redis.aclose()
    jwt = encode_jwt(user_id=user_id, session_id=session_id, tenant_id=tenant_id)
    return user_id, jwt


async def _admin(dsn: str, redis_url: str, *, tenant: UUID, slug: str) -> str:
    _id, jwt = await _seed_user_with_jwt(
        dsn, redis_url, tenant_id=tenant, email=f"admin@{slug}.example.com", role="tenant_admin"
    )
    return jwt


async def _seed_project(dsn: str, *, tenant: UUID, name: str) -> UUID:
    project_id = uuid4()
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO projects (id, tenant_id, name, status) VALUES ($1, $2, $3, 'active')",
            project_id,
            tenant,
            name,
        )
    finally:
        await conn.close()
    return project_id


async def _seed_plan(dsn: str, *, tenant: UUID, project_id: UUID, title: str) -> UUID:
    plan_id = uuid4()
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO plans (id, tenant_id, project_id, title, created_by) "
            "VALUES ($1, $2, $3, $4, NULL)",
            plan_id,
            tenant,
            project_id,
            title,
        )
    finally:
        await conn.close()
    return plan_id


async def _seed_agent(dsn: str, *, tenant: UUID, name: str, role: str) -> UUID:
    agent_id = uuid4()
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO agents (id, tenant_id, name, role, system_prompt, scope) "
            "VALUES ($1, $2, $3, $4, 'be helpful', 'global_tenant_template')",
            agent_id,
            tenant,
            name,
            role,
        )
    finally:
        await conn.close()
    return agent_id


async def _seed_task(
    dsn: str,
    *,
    tenant: UUID,
    project_id: UUID,
    plan_id: UUID | None,
    title: str,
    retry_count: int = 0,
) -> UUID:
    task_id = uuid4()
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO tasks (id, tenant_id, project_id, plan_id, title, status, retry_count) "
            "VALUES ($1, $2, $3, $4, $5, 'done', $6)",
            task_id,
            tenant,
            project_id,
            plan_id,
            title,
            retry_count,
        )
    finally:
        await conn.close()
    return task_id


def _model_steps(
    model: str, *, tokens_in: int, tokens_out: int, cost: float, secret: str = "TOP-SECRET-PROMPT"
) -> list[dict]:
    """A steps_log with one model_call step carrying a (would-be) secret prompt.

    The ``prompt`` / ``completion`` fields stand in for raw model traffic; the
    export must NEVER surface them (the no-PII assertion checks for ``secret``).
    """
    return [
        {
            "index": 0,
            "kind": "model_call",
            "node": "act",
            "status": "ok",
            "model": model,
            "tokens_in": tokens_in,
            "tokens_out": tokens_out,
            "total_tokens": tokens_in + tokens_out,
            "cost_usd": cost,
            "prompt": secret,
            "completion": secret,
        }
    ]


async def _seed_execution(
    dsn: str,
    *,
    tenant: UUID,
    task_id: UUID,
    agent_id: UUID | None,
    status: str = "done",
    total_tokens: int = 0,
    total_cost_usd: float = 0.0,
    duration_ms: int | None = None,
    steps_log: list[dict] | None = None,
    created_at: datetime | None = None,
) -> UUID:
    execution_id = uuid4()
    now = created_at or datetime.now(tz=UTC)
    started = now
    completed = now + timedelta(milliseconds=duration_ms) if duration_ms is not None else None
    # prod-13 task_prod13_18 / migración 0139: ver la nota gemela en
    # `test_tenant_stats_dashboard.py`. La siembra hace de escritor y deriva la
    # proyección con la MISMA función que `execution_repo`.
    rollup = steps_rollup(steps_log or [])
    # `executions` está particionada por mes y SIN partición DEFAULT (ADR 0151): una
    # fila retrofechada muere en el propio INSERT si su mes no tiene partición. Aquí
    # el llamante del CSV siembra `now - 1h`, que cae en el mes anterior durante la
    # primera hora del mes. Ver
    # docs/03-guides/gotchas/sembrar-filas-retrofechadas-en-tabla-particionada.md
    await ensure_partition_for(dsn, "executions", now)
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "INSERT INTO executions "
            "(id, tenant_id, task_id, agent_id, status, steps_log, total_tokens, "
            " total_cost_usd, started_at, completed_at, created_at,"
            " last_model, tokens_in, tokens_out) "
            "VALUES ($1, $2, $3, $4, $5, $6::jsonb, $7, $8, $9, $10, $11, $12, $13, $14)",
            execution_id,
            tenant,
            task_id,
            agent_id,
            status,
            json.dumps(steps_log or []),
            total_tokens,
            total_cost_usd,
            started,
            completed,
            now,
            rollup.last_model,
            rollup.tokens_in,
            rollup.tokens_out,
        )
    finally:
        await conn.close()
    return execution_id


async def _truncate_all(dsn: str) -> None:
    conn = await asyncpg.connect(dsn)
    try:
        await conn.execute(
            "TRUNCATE executions, tasks, plans, agents, projects, "
            "user_org_memberships, organizations, users RESTART IDENTITY CASCADE"
        )
    finally:
        await conn.close()


# ---------------------------------------------------------------------------
# App fixture: real api-server wired to the test DB + Redis
# ---------------------------------------------------------------------------
@pytest.fixture()
def configured_app(
    alembic_config,
    app_database_url: str,
    admin_database_url: str,
    test_redis_url: str,
    monkeypatch: pytest.MonkeyPatch,
):
    command.upgrade(alembic_config, "head")

    from tests.integration.conftest import _flush_redis, _grant_app_user_existing_tables

    asyncio.run(_grant_app_user_existing_tables())
    asyncio.run(_flush_redis(test_redis_url))

    monkeypatch.setenv("API_SERVER_DATABASE_URL", app_database_url)
    monkeypatch.setenv("API_SERVER_ADMIN_DATABASE_URL", admin_database_url)
    monkeypatch.setenv("API_SERVER_REDIS_URL", test_redis_url)
    monkeypatch.setenv("API_SERVER_JWT_SECRET", "test-secret")
    monkeypatch.setenv("API_SERVER_SSO_ENCRYPTION_KEY", "test-sso-encryption-key")
    monkeypatch.setenv("API_SERVER_SSO_REDIRECT_BASE_URL", "http://testserver")
    monkeypatch.delenv("API_SERVER_VAULT_TOKEN", raising=False)

    from api_server.auth.deps import reset_redis_cache
    from api_server.config import get_settings
    from api_server.db.session import reset_engine_cache

    get_settings.cache_clear()
    reset_engine_cache()
    reset_redis_cache()

    from api_server.main import create_app

    app = create_app()
    try:
        yield app
    finally:
        app.dependency_overrides.clear()
        reset_engine_cache()
        reset_redis_cache()
        get_settings.cache_clear()


def _auth(jwt: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {jwt}"}


def _client(app: object) -> AsyncClient:
    return AsyncClient(transport=ASGITransport(app=app), base_url="http://testserver")


async def _seed_two_runs(dsn: str, *, tenant: UUID) -> tuple[UUID, UUID, UUID, UUID]:
    """Seed a project/plan/two agents/task + two executions; return their ids."""
    project = await _seed_project(dsn, tenant=tenant, name="P")
    plan = await _seed_plan(dsn, tenant=tenant, project_id=project, title="Plan A")
    agent_be = await _seed_agent(dsn, tenant=tenant, name="Backend", role="backend")
    agent_fe = await _seed_agent(dsn, tenant=tenant, name="Frontend", role="frontend")
    task = await _seed_task(
        dsn, tenant=tenant, project_id=project, plan_id=plan, title="Build login", retry_count=2
    )
    ex_be = await _seed_execution(
        dsn,
        tenant=tenant,
        task_id=task,
        agent_id=agent_be,
        status="done",
        total_tokens=1500,
        total_cost_usd=0.05,
        duration_ms=1234,
        steps_log=_model_steps("claude-opus", tokens_in=1000, tokens_out=500, cost=0.05),
    )
    ex_fe = await _seed_execution(
        dsn,
        tenant=tenant,
        task_id=task,
        agent_id=agent_fe,
        status="aborted",
        total_tokens=200,
        total_cost_usd=0.001,
        duration_ms=None,
        steps_log=_model_steps("claude-haiku", tokens_in=150, tokens_out=50, cost=0.001),
    )
    return plan, agent_be, ex_be, ex_fe


# ---------------------------------------------------------------------------
# CSV: right header + rows for the tenant's data
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_csv_export_header_and_rows(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="acme")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="acme")
    _plan, _agent_be, ex_be, ex_fe = await _seed_two_runs(migrations_pg_dsn, tenant=tenant)

    async with _client(configured_app) as client:
        resp = await client.get("/tenant-stats/runs/export?format=csv", headers=_auth(jwt))
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"].startswith("text/csv")
        assert "attachment" in resp.headers["content-disposition"]
        assert ".csv" in resp.headers["content-disposition"]

        # Parse it back: 1 header + 2 rows.
        text = resp.content.decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(text)))
        header = reader[0]
        assert header[0] == "execution_id"
        assert "total_cost_usd" in header
        assert "model" in header
        assert "verdict" in header
        data_rows = reader[1:]
        assert len(data_rows) == 2

        by_id = {r[header.index("execution_id")]: r for r in data_rows}
        assert str(ex_be) in by_id
        assert str(ex_fe) in by_id
        be = by_id[str(ex_be)]
        assert be[header.index("plan_title")] == "Plan A"
        assert be[header.index("task_title")] == "Build login"
        assert be[header.index("agent_name")] == "Backend"
        assert be[header.index("agent_role")] == "backend"
        assert be[header.index("model")] == "claude-opus"
        assert be[header.index("verdict")] == "done"
        assert be[header.index("succeeded")] == "true"
        assert be[header.index("retry_count")] == "2"
        assert be[header.index("duration_ms")] == "1234"
        assert be[header.index("total_cost_usd")] == "0.050000"

        # No raw model traffic (prompt/completion) leaked into the CSV.
        assert "TOP-SECRET-PROMPT" not in text


# ---------------------------------------------------------------------------
# XLSX: a valid file with header + rows
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_xlsx_export_valid_file(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    from openpyxl import load_workbook

    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="acme")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="acme")
    await _seed_two_runs(migrations_pg_dsn, tenant=tenant)

    async with _client(configured_app) as client:
        resp = await client.get("/tenant-stats/runs/export?format=xlsx", headers=_auth(jwt))
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in resp.headers["content-disposition"]
        # A real XLSX is a zip: starts with the PK magic bytes.
        assert resp.content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Runs"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "execution_id"
        # header + 2 data rows
        assert len(rows) == 3
        # No secret prompt anywhere in the workbook cells.
        flat = " ".join(str(c) for r in rows for c in r if c is not None)
        assert "TOP-SECRET-PROMPT" not in flat


# ---------------------------------------------------------------------------
# PDF → print-ready HTML document
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_pdf_export_is_printable_html(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="acme")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="acme")
    await _seed_two_runs(migrations_pg_dsn, tenant=tenant)

    async with _client(configured_app) as client:
        resp = await client.get("/tenant-stats/runs/export?format=pdf", headers=_auth(jwt))
        assert resp.status_code == 200, resp.text
        # Degraded honestly: a print-ready HTML document, not a binary PDF.
        assert resp.headers["content-type"].startswith("text/html")
        body = resp.content.decode("utf-8")
        assert body.lstrip().lower().startswith("<!doctype html>")
        assert "<table" in body
        assert "Build login" in body
        # Header + the consumption summary block are present.
        assert "Tenant runs export" in body
        assert "Consumption" in body
        assert "TOP-SECRET-PROMPT" not in body


# ---------------------------------------------------------------------------
# Filters narrow the export
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_export_respects_filters(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="acme")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="acme")
    _plan, agent_be, ex_be, _ex_fe = await _seed_two_runs(migrations_pg_dsn, tenant=tenant)

    async with _client(configured_app) as client:
        resp = await client.get(
            f"/tenant-stats/runs/export?format=csv&agent_id={agent_be}", headers=_auth(jwt)
        )
        assert resp.status_code == 200, resp.text
        reader = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
        data_rows = reader[1:]
        assert len(data_rows) == 1
        assert data_rows[0][reader[0].index("execution_id")] == str(ex_be)


# ---------------------------------------------------------------------------
# Bad format → 422 (the enum rejects unknown values)
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_unknown_format_is_422(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="acme")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="acme")

    async with _client(configured_app) as client:
        resp = await client.get("/tenant-stats/runs/export?format=docx", headers=_auth(jwt))
        assert resp.status_code == 422, resp.text


# ---------------------------------------------------------------------------
# RBAC: a plain member (tenant_user) is denied (403)
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_plain_member_denied_403(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="acme")
    _uid, member_jwt = await _seed_user_with_jwt(
        migrations_pg_dsn,
        test_redis_url,
        tenant_id=tenant,
        email="member@acme.example.com",
        role="tenant_user",
    )

    async with _client(configured_app) as client:
        for fmt in ("csv", "xlsx", "pdf"):
            resp = await client.get(
                f"/tenant-stats/runs/export?format={fmt}", headers=_auth(member_jwt)
            )
            assert resp.status_code == 403, (fmt, resp.text)


# ---------------------------------------------------------------------------
# Cross-tenant: tenant A's export never contains tenant B's rows
# ---------------------------------------------------------------------------
@pytest.mark.cross_tenant
@pytest.mark.asyncio
async def test_cross_tenant_isolation(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    await _truncate_all(migrations_pg_dsn)
    tenant_a = await _seed_tenant(migrations_pg_dsn, slug="alpha")
    tenant_b = await _seed_tenant(migrations_pg_dsn, slug="bravo")
    jwt_a = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant_a, slug="alpha")

    # B owns a project + agent + task + a costly execution with a secret prompt.
    project_b = await _seed_project(migrations_pg_dsn, tenant=tenant_b, name="PB")
    agent_b = await _seed_agent(migrations_pg_dsn, tenant=tenant_b, name="B-agent", role="backend")
    task_b = await _seed_task(
        migrations_pg_dsn,
        tenant=tenant_b,
        project_id=project_b,
        plan_id=None,
        title="B-secret-task",
    )
    ex_b = await _seed_execution(
        migrations_pg_dsn,
        tenant=tenant_b,
        task_id=task_b,
        agent_id=agent_b,
        status="done",
        total_tokens=9999,
        total_cost_usd=1.23,
        duration_ms=4000,
        steps_log=_model_steps("b-model", tokens_in=8000, tokens_out=1999, cost=1.23),
    )

    async with _client(configured_app) as client:
        # CSV
        csv_resp = await client.get("/tenant-stats/runs/export?format=csv", headers=_auth(jwt_a))
        assert csv_resp.status_code == 200, csv_resp.text
        csv_text = csv_resp.content.decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(csv_text)))
        assert reader[1:] == []  # header only, no B rows
        assert str(ex_b) not in csv_text
        assert "B-secret-task" not in csv_text

        # PDF/HTML
        pdf_resp = await client.get("/tenant-stats/runs/export?format=pdf", headers=_auth(jwt_a))
        assert pdf_resp.status_code == 200, pdf_resp.text
        pdf_body = pdf_resp.content.decode("utf-8")
        assert "B-secret-task" not in pdf_body
        assert str(ex_b) not in pdf_body


# ---------------------------------------------------------------------------
# prod-13 task_prod13_18 — el export pagina por KEYSET, no por offset
# ---------------------------------------------------------------------------
@pytest.mark.asyncio
async def test_export_resumes_past_the_row_cap_with_the_keyset_cursor(
    configured_app, migrations_pg_dsn: str, test_redis_url: str, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Un tenant con más runs que el tope se los baja en ficheros consecutivos.

    Antes, todo lo que pasase de ``MAX_EXPORT_ROWS`` se perdía **en silencio**:
    el export devolvía las N primeras filas y nada decía que hubiera más. Ahora
    la respuesta lleva ``X-Next-Cursor`` cuando llenó la página y el mismo token
    reanuda el export justo después de la última fila entregada.

    El tope se baja a 2 con monkeypatch en vez de sembrar 5.001 ejecuciones: lo
    que se prueba es el CONTRATO de reanudación, y sembrar cinco mil filas para
    eso cambiaría un test de segundos por uno de minutos sin probar nada más.

    Y se comprueba lo que un cursor mal hecho rompe: que la segunda página no
    repite ni se salta filas — la razón por la que el keyset compara la FILA
    ``(created_at, id)`` y no dos predicados sueltos.
    """
    import api_server.routers.tenant_stats as stats_router

    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="keyset")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="keyset")
    project = await _seed_project(migrations_pg_dsn, tenant=tenant, name="P")
    task = await _seed_task(
        migrations_pg_dsn, tenant=tenant, project_id=project, plan_id=None, title="T"
    )
    base = datetime.now(tz=UTC) - timedelta(hours=1)
    seeded = [
        await _seed_execution(
            migrations_pg_dsn,
            tenant=tenant,
            task_id=task,
            agent_id=None,
            created_at=base + timedelta(minutes=i),
        )
        for i in range(5)
    ]

    monkeypatch.setattr(stats_router, "MAX_EXPORT_ROWS", 2)

    def _ids(body: bytes) -> list[str]:
        reader = list(csv.reader(io.StringIO(body.decode("utf-8-sig"))))
        column = reader[0].index("execution_id")
        return [row[column] for row in reader[1:]]

    collected: list[str] = []
    async with _client(configured_app) as client:
        url = "/tenant-stats/runs/export?format=csv"
        for _page in range(4):
            resp = await client.get(url, headers=_auth(jwt))
            assert resp.status_code == 200, resp.text
            collected.extend(_ids(resp.content))
            cursor = resp.headers.get("X-Next-Cursor")
            if cursor is None:
                break
            url = f"/tenant-stats/runs/export?format=csv&cursor={cursor}"
        else:  # pragma: no cover - sólo si el cursor no termina nunca
            raise AssertionError("el export nunca dejó de ofrecer X-Next-Cursor")

    # Las cinco, una sola vez cada una y en orden descendente por created_at.
    assert collected == [
        str(x) for x in reversed(seeded)
    ], "la paginación del export repite o se salta filas"
    assert len(set(collected)) == len(collected)


@pytest.mark.asyncio
async def test_a_corrupt_export_cursor_is_a_400_not_a_500(
    configured_app, migrations_pg_dsn: str, test_redis_url: str
) -> None:
    """El cursor lo manda el cliente: uno roto es un error suyo. Un 500 lo
    convertiría en una alerta de servidor y en ruido de guardia."""
    await _truncate_all(migrations_pg_dsn)
    tenant = await _seed_tenant(migrations_pg_dsn, slug="badcur")
    jwt = await _admin(migrations_pg_dsn, test_redis_url, tenant=tenant, slug="badcur")

    async with _client(configured_app) as client:
        resp = await client.get(
            "/tenant-stats/runs/export?format=csv&cursor=not-a-real-cursor", headers=_auth(jwt)
        )
    assert resp.status_code == 400
